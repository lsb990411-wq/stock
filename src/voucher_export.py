"""회계 프로그램 '엑셀자료일반전표전송' 원본 양식에 매매분개 채우기."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from .fifo import sells_by_trade_id
from .models import AccountConfig, SellResult, Trade

# 기본 계정과목 (코드는 AccountConfig로 덮어씀, 이름은 고정)
_ACCT_NAMES = {
    "security": "투자유가증권",
    "fee": "지급수수료",
    "deposit": "기타제예금",
    "gain": "투자자산처분이익",
    "loss": "투자자산처분손실",
}

DR, CR = 3, 4  # 차변, 대변

# 원본 xls: 헤더=10행, 데이터=11행부터, 컬럼 28개
HEADER_ROW = 10
DATA_START_ROW = 11
NUM_COLUMNS = 28

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"
TEMPLATE_XLS = TEMPLATE_DIR / "엑셀자료일반전표전송.xls"
TEMPLATE_XLSX = TEMPLATE_DIR / "엑셀자료일반전표전송.xlsx"


@dataclass(frozen=True)
class VoucherAccounts:
    security: tuple[str, str]
    fee: tuple[str, str]
    deposit: tuple[str, str]
    gain: tuple[str, str]
    loss: tuple[str, str]

    @classmethod
    def from_config(cls, config: AccountConfig | None = None) -> VoucherAccounts:
        cfg = (config or AccountConfig()).normalize()
        return cls(
            security=(cfg.security_code, _ACCT_NAMES["security"]),
            fee=(cfg.fee_code, _ACCT_NAMES["fee"]),
            deposit=(cfg.deposit_code, _ACCT_NAMES["deposit"]),
            gain=(cfg.gain_code, _ACCT_NAMES["gain"]),
            loss=(cfg.loss_code, _ACCT_NAMES["loss"]),
        )


@dataclass
class VoucherLine:
    ymd: str
    side: int  # 3=차변, 4=대변
    acct_code: str
    acct_name: str
    partner_code: str
    partner_name: str
    summary_code: str
    summary: str
    amount: int

    def as_row(self) -> list:
        """원본 양식 28컬럼 순서에 맞춤 (값 없는 칸은 공란)."""
        row = [""] * NUM_COLUMNS
        row[0] = self.ymd
        row[1] = self.side
        row[2] = self.acct_code
        row[3] = self.acct_name
        row[4] = self.partner_code
        row[5] = (self.partner_name or "")[:30]
        row[6] = self.summary_code
        row[7] = self.summary
        row[8] = self.amount
        return row


def _ymd(trade_date: str) -> str:
    return "".join(ch for ch in str(trade_date) if ch.isdigit())[:8]


def _won(value: float | int | None) -> int:
    if value is None:
        return 0
    return int(round(float(value)))


def _fmt_num(value: float | int | None) -> str:
    """적요용 천단위 구분 숫자(원화·정수)."""
    return f"{_won(value):,}"


def _fmt_qty_overseas(qty: float | int | None) -> str:
    """해외 수량: 정수면 정수, 아니면 소수(끝 0 제거)."""
    q = float(qty or 0)
    if abs(q - round(q)) < 1e-9:
        return f"{int(round(q))}"
    return f"{q:.4f}".rstrip("0").rstrip(".")


def _fmt_fx_price(value: float | int | None) -> str:
    """외화 단가: 최대 소수 4자리, 끝 0 제거."""
    s = f"{float(value or 0):.4f}".rstrip("0").rstrip(".")
    return s if s else "0"


def _fmt_fx_amount(value: float | int | None) -> str:
    """외화 지급액: 소수 2자리."""
    return f"{float(value or 0):.2f}"


def _fmt_fx_rate(value: float | int | None) -> str:
    """적용 환율: 소수 2자리 (예시: 1470.50)."""
    return f"{float(value or 0):.2f}"


def _fmt_num_plain(value: float | int | None) -> str:
    """적요용 숫자(천단위 없이, 끝 0 제거)."""
    x = float(value or 0)
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.6f}".rstrip("0").rstrip(".")


def _is_overseas_trade(trade: Trade) -> bool:
    if getattr(trade, "is_overseas", False):
        return True
    market = str(getattr(trade, "stock_market", "") or "")
    if market.lower() in {"overseas", "해외", "해외주식"}:
        return True
    if float(getattr(trade, "price_fx", 0) or 0) > 0:
        return True
    return float(getattr(trade, "fx_rate", 0) or 0) > 0


def _trade_side_label(trade: Trade) -> str:
    side = str(trade.side or "").upper()
    if side == "BUY":
        return "매수"
    if side == "SELL":
        return "매도"
    if side == "DIVIDEND":
        return "배당"
    return str(trade.side or "거래")


def _trade_fx_parts(trade: Trade) -> tuple[float, float, float, str]:
    """수량, 외화단가, 환율, 통화."""
    qty = float(trade.quantity or 0)
    price_fx = float(getattr(trade, "price_fx", 0) or 0)
    fx = float(getattr(trade, "fx_rate", 0) or 0)
    ccy = str(getattr(trade, "currency", "") or "").strip().upper() or "USD"
    price_krw = float(trade.price or 0)
    if price_fx <= 0 and fx > 0 and price_krw > 0 and ccy != "KRW":
        price_fx = price_krw / fx
    return qty, price_fx, fx, ccy


def _is_kb_trade(trade: Trade) -> bool:
    """KB증권 거래 여부 (해외 전표 적요 분기용)."""
    src = str(getattr(trade, "source", "") or "").lower()
    acct = str(getattr(trade, "account_name", "") or "")
    if "kb" in src:
        return True
    return "KB" in acct.upper() or "KB증권" in acct


def _overseas_settle_fx(trade: Trade) -> float:
    """외화정산금액 ≈ 거래금액 ± 외화수수료·제세금.

    매수: 수량×단가 + 수수료 + 제세금
    매도: 수량×단가 − 수수료 − 제세금
    """
    qty, price_fx, fx, _ccy = _trade_fx_parts(trade)
    fee_fx = float(getattr(trade, "fee_fx", 0) or 0)
    tax_fx = float(getattr(trade, "tax_fx", 0) or 0)
    if fee_fx <= 0 and fx > 0 and float(trade.fee or 0) > 0:
        fee_fx = float(trade.fee) / fx
    if tax_fx <= 0 and fx > 0 and float(getattr(trade, "tax", 0) or 0) > 0:
        tax_fx = float(trade.tax) / fx
    gross = abs(qty * price_fx)
    side = str(trade.side or "").upper()
    if side == "BUY":
        return gross + abs(fee_fx) + abs(tax_fx)
    if side == "SELL":
        return max(0.0, gross - abs(fee_fx) - abs(tax_fx))
    return gross


def build_overseas_remark(trade: Trade) -> str:
    """해외주식 전표 적요 (옵션1: 종목명 / @수량 * 단가 * 환율).

    예: INVESCO NASDAQ 100 매수 / @40주 * $247.67 * 1442.00원
    """
    name = (trade.stock_name or trade.stock_code or "").strip()
    qty, price_fx, fx, ccy = _trade_fx_parts(trade)
    side = str(trade.side or "").upper()
    label = _trade_side_label(trade)

    if side == "DIVIDEND":
        amount_usd = price_fx * (qty if qty > 0 else 1.0)
        return (
            f"{name} 배당입금 ${_fmt_fx_amount(amount_usd)}"
            f" * {_fmt_fx_rate(fx)}원"
        )

    if price_fx > 0 and fx > 0:
        if ccy == "USD":
            px = f"${_fmt_fx_price(price_fx)}"
        else:
            px = f"{ccy} {_fmt_fx_price(price_fx)}"
        return (
            f"{name} {label} / @{_fmt_qty_overseas(qty)}주"
            f" * {px} * {_fmt_fx_rate(fx)}원"
        )
    return f"{name} {label} / @{_fmt_qty_overseas(qty)}주"


def build_overseas_remark_amount(
    trade: Trade,
    *,
    use_settlement_fx: bool = False,
) -> str:
    """해외주식 전표 적요 (옵션2: 통화 외화총액 / 수량*단가 * 환율).

    예: USD 1862.29 / 5주*371.529 * 1528.6
    use_settlement_fx=True 이면 앞부분 금액을 외화정산금액(거래금액±수수료)으로 표기.
    """
    qty, price_fx, fx, ccy = _trade_fx_parts(trade)
    side = str(trade.side or "").upper()

    if side == "DIVIDEND":
        amount_usd = price_fx * (qty if qty > 0 else 1.0)
        body = f"{ccy} {_fmt_num_plain(amount_usd)}"
        if fx > 0:
            body += f" * {_fmt_num_plain(fx)}"
        return body

    if price_fx > 0 and qty > 0:
        fx_amt = (
            _overseas_settle_fx(trade) if use_settlement_fx else qty * price_fx
        )
        body = (
            f"{ccy} {_fmt_num_plain(fx_amt)} / "
            f"{_fmt_qty_overseas(qty)}주*{_fmt_fx_price(price_fx)}"
        )
        if fx > 0:
            body += f" * {_fmt_num_plain(fx)}"
        return body

    name = (trade.stock_name or trade.stock_code or "").strip()
    return f"{name} {_trade_side_label(trade)}"


def _overseas_fee_tax_fx(trade: Trade) -> tuple[float, float, float, str]:
    """외화수수료, 외화제세금, 환율, 통화. 외화 없으면 원화÷환율로 역산."""
    _qty, _price_fx, fx, ccy = _trade_fx_parts(trade)
    fee_fx = float(getattr(trade, "fee_fx", 0) or 0)
    tax_fx = float(getattr(trade, "tax_fx", 0) or 0)
    if fee_fx <= 0 and fx > 0 and float(trade.fee or 0) > 0:
        fee_fx = float(trade.fee) / fx
    if tax_fx <= 0 and fx > 0 and float(getattr(trade, "tax", 0) or 0) > 0:
        tax_fx = float(trade.tax) / fx
    return abs(fee_fx), abs(tax_fx), fx, ccy


def build_overseas_fee_remark(
    trade: Trade,
    *,
    kind: str = "fee",
    remark_mode: str = "stock",
) -> str:
    """해외주식 수수료·제세금 적요 (외화·환율 포함).

    amount: USD 0.69 * 1473.5
    stock:  주식매수수수료 / $0.69 * 1473.50원
    """
    fee_fx, tax_fx, fx, ccy = _overseas_fee_tax_fx(trade)
    amt_fx = tax_fx if kind == "tax" else fee_fx
    if amt_fx <= 0:
        return "주식매도제세금" if kind == "tax" else (
            "주식매도수수료" if str(trade.side or "").upper() == "SELL" else "주식매수수수료"
        )

    side = str(trade.side or "").upper()
    if kind == "tax":
        label = "주식매도제세금"
    elif side == "SELL":
        label = "주식매도수수료"
    else:
        label = "주식매수수수료"

    if remark_mode == "amount":
        body = f"{ccy} {_fmt_num_plain(amt_fx)}"
        if fx > 0:
            body += f" * {_fmt_num_plain(fx)}"
        return body

    if ccy == "USD":
        px = f"${_fmt_fx_amount(amt_fx)}"
    else:
        px = f"{ccy} {_fmt_fx_amount(amt_fx)}"
    if fx > 0:
        return f"{label} / {px} * {_fmt_fx_rate(fx)}원"
    return f"{label} / {px}"


def _fee_summary(trade: Trade, remark_mode: str, *, kind: str = "fee") -> str:
    """수수료·제세금 적요. 해외면 외화·환율 표기."""
    if _is_overseas_trade(trade):
        return build_overseas_fee_remark(
            trade, kind=kind, remark_mode=remark_mode
        )
    if kind == "tax":
        return "주식매도제세금"
    return (
        "주식매도수수료"
        if str(trade.side or "").upper() == "SELL"
        else "주식매수수수료"
    )


def _deposit_summary(trade: Trade, remark_mode: str, base_memo: str) -> str:
    """기타제예금 적요.

    KB증권 + 옵션2(amount): 거래금액 대신 외화정산금액 사용.
    """
    if (
        remark_mode == "amount"
        and _is_overseas_trade(trade)
        and _is_kb_trade(trade)
    ):
        return build_overseas_remark_amount(trade, use_settlement_fx=True)
    return base_memo


def _pad_remark_by_slash(texts: list[str]) -> list[str]:
    """옵션2 적요: '/' 앞부분만 최장 길이에 오른쪽 공백 패딩 후 재조합.

    `앞부분.ljust(max_len) + " / " + 뒷부분`
    """
    cleaned = ["" if t is None else str(t) for t in texts]
    if not cleaned:
        return cleaned

    fronts: list[str] = []
    backs: list[str] = []
    has_sep: list[bool] = []
    for m in cleaned:
        if "/" in m:
            left, right = m.split("/", 1)
            fronts.append(left.rstrip())
            backs.append(right.lstrip())
            has_sep.append(True)
        else:
            fronts.append(m)
            backs.append("")
            has_sep.append(False)

    sep_lens = [len(f) for f, sep in zip(fronts, has_sep) if sep]
    if not sep_lens:
        return cleaned
    max_len = max(sep_lens)

    out: list[str] = []
    for front, back, sep in zip(fronts, backs, has_sep):
        if not sep:
            out.append(front)
            continue
        out.append(front.ljust(max_len) + " / " + back)
    return out


def _pad_summaries(lines: list[VoucherLine]) -> list[VoucherLine]:
    """옵션2: '/' 포함 거래 적요의 앞부분만 공백 패딩."""
    if not lines:
        return lines
    idxs = [i for i, ln in enumerate(lines) if "/" in (ln.summary or "")]
    if not idxs:
        return lines
    padded = _pad_remark_by_slash([lines[i].summary or "" for i in idxs])
    for i, text in zip(idxs, padded):
        lines[i].summary = text
    return lines


def _partner_for_trade(
    trade: Trade,
    partner_by_stock_id: Mapping[int, str] | None,
) -> str:
    """종목별 회계 거래처코드. 없으면 종목코드(ticker) 사용."""
    mapped = ""
    if partner_by_stock_id and trade.stock_id in partner_by_stock_id:
        mapped = (partner_by_stock_id.get(trade.stock_id) or "").strip()
    if mapped:
        return mapped
    return (trade.stock_code or "").strip()


def _line(
    *,
    ymd: str,
    side: int,
    acct: tuple[str, str],
    partner_code: str,
    partner_name: str,
    amount: int,
    summary: str,
    summary_code: str = "",
) -> VoucherLine | None:
    if amount == 0:
        return None
    return VoucherLine(
        ymd=ymd,
        side=side,
        acct_code=acct[0],
        acct_name=acct[1],
        partner_code=partner_code or "",
        partner_name=partner_name or "",
        summary_code=summary_code,
        summary=summary,
        amount=abs(amount),
    )


def _trade_summary_buy(trade: Trade, remark_mode: str = "stock") -> str:
    if _is_overseas_trade(trade):
        if remark_mode == "amount":
            return build_overseas_remark_amount(trade)
        return build_overseas_remark(trade)
    name = (trade.stock_name or trade.stock_code or "").strip()
    if remark_mode == "amount":
        qty = float(trade.quantity or 0)
        price = float(trade.price or 0)
        return (
            f"KRW {_fmt_num_plain(qty * price)} / "
            f"{_fmt_qty_overseas(qty)}주*{_fmt_num_plain(price)}"
        )
    return (
        f"{name} 매수 / @{_fmt_qty_overseas(trade.quantity)}주"
        f" * {_fmt_num(trade.price)}원"
    )


def _trade_summary_sell(trade: Trade, remark_mode: str = "stock") -> str:
    if _is_overseas_trade(trade):
        if remark_mode == "amount":
            return build_overseas_remark_amount(trade)
        return build_overseas_remark(trade)
    name = (trade.stock_name or trade.stock_code or "").strip()
    if remark_mode == "amount":
        qty = float(trade.quantity or 0)
        price = float(trade.price or 0)
        return (
            f"KRW {_fmt_num_plain(qty * price)} / "
            f"{_fmt_qty_overseas(qty)}주*{_fmt_num_plain(price)}"
        )
    return (
        f"{name} 매도 / @{_fmt_qty_overseas(trade.quantity)}주"
        f" * {_fmt_num(trade.price)}원"
    )


def _trade_summary_dividend(trade: Trade, remark_mode: str = "stock") -> str:
    if _is_overseas_trade(trade):
        if remark_mode == "amount":
            return build_overseas_remark_amount(trade)
        return build_overseas_remark(trade)
    name = (trade.stock_name or trade.stock_code or "").strip()
    return f"{name} 배당입금 {_fmt_num(trade.settlement_amount or trade.price)}"



def _buy_lines(
    trade: Trade,
    accounts: VoucherAccounts,
    partner_code: str,
    remark_mode: str = "stock",
) -> list[VoucherLine]:
    ymd = _ymd(trade.trade_date)
    qty = trade.quantity
    price = trade.price
    stock_name = (trade.stock_name or "").strip()
    broker_code, broker_name = "", ""

    principal = _won(qty * price)
    fee = _won(trade.fee)
    tax = _won(getattr(trade, "tax", 0) or 0)
    if trade.settlement_amount is not None:
        total_out = _won(trade.settlement_amount)
    else:
        total_out = principal + fee + tax

    memo = _trade_summary_buy(trade, remark_mode=remark_mode)
    deposit_memo = _deposit_summary(trade, remark_mode, memo)
    fee_memo = _fee_summary(trade, remark_mode, kind="fee")
    lines: list[VoucherLine] = []
    for item in (
        _line(
            ymd=ymd,
            side=DR,
            acct=accounts.security,
            partner_code=partner_code,
            partner_name=stock_name,
            amount=principal,
            summary=memo,
        ),
        _line(
            ymd=ymd,
            side=DR,
            acct=accounts.fee,
            partner_code=partner_code,
            partner_name=stock_name,
            amount=fee,
            summary=fee_memo,
        ),
        _line(
            ymd=ymd,
            side=CR,
            acct=accounts.deposit,
            partner_code=broker_code,
            partner_name=broker_name,
            amount=total_out,
            summary=deposit_memo,
        ),
    ):
        if item:
            lines.append(item)
    return lines


def _sell_lines(
    trade: Trade,
    sell: SellResult | None,
    accounts: VoucherAccounts,
    partner_code: str,
    remark_mode: str = "stock",
) -> list[VoucherLine]:
    """매도 분개: 투자유가증권 대변을 FIFO 매수 레이어별로 분할."""
    ymd = _ymd(trade.trade_date)
    qty = trade.quantity
    price = trade.price
    stock_name = (trade.stock_name or "").strip()
    broker_code, broker_name = "", ""

    lines: list[VoucherLine] = []
    book = 0

    # 1) FIFO 소진 레이어별 투자유가증권 대변
    if sell and sell.matches:
        for match in sell.matches:
            used_qty = match.matched_qty
            layer_price = match.buy_price
            layer_amt = _won(used_qty * layer_price)
            book += layer_amt
            item = _line(
                ymd=ymd,
                side=CR,
                acct=accounts.security,
                partner_code=partner_code,
                partner_name=stock_name,
                amount=layer_amt,
                summary=(
                    f"주식매도 원가 @{_fmt_num(used_qty)} * {_fmt_num(layer_price)}"
                ),
            )
            if item:
                lines.append(item)

    fee = _won(trade.fee)
    tax = _won(getattr(trade, "tax", 0) or 0)
    gross = _won(qty * price)
    if trade.settlement_amount is not None:
        bank_in = _won(trade.settlement_amount)
    else:
        bank_in = max(gross - fee - tax, 0)

    sell_memo = _trade_summary_sell(trade, remark_mode=remark_mode)
    deposit_memo = _deposit_summary(trade, remark_mode, sell_memo)
    fee_memo = _fee_summary(trade, remark_mode, kind="fee")
    tax_memo = _fee_summary(trade, remark_mode, kind="tax")
    # 차대 균형: 입금 + 매도비용 - FIFO 원가
    pnl = bank_in + fee + tax - book

    # 2) 기타제예금 / 수수료 / 제세금 / 처분손익
    for item in (
        _line(
            ymd=ymd,
            side=DR,
            acct=accounts.deposit,
            partner_code=broker_code,
            partner_name=broker_name,
            amount=bank_in,
            summary=deposit_memo,
        ),
        _line(
            ymd=ymd,
            side=DR,
            acct=accounts.fee,
            partner_code=broker_code,
            partner_name=broker_name,
            amount=fee,
            summary=fee_memo,
        ),
        _line(
            ymd=ymd,
            side=DR,
            acct=accounts.fee,
            partner_code=broker_code,
            partner_name=broker_name,
            amount=tax,
            summary=tax_memo,
        ),
    ):
        if item:
            lines.append(item)

    if pnl > 0:
        gain = _line(
            ymd=ymd,
            side=CR,
            acct=accounts.gain,
            partner_code=partner_code,
            partner_name=stock_name,
            amount=pnl,
            summary="주식매도",
        )
        if gain:
            lines.append(gain)
    elif pnl < 0:
        loss = _line(
            ymd=ymd,
            side=DR,
            acct=accounts.loss,
            partner_code=partner_code,
            partner_name=stock_name,
            amount=abs(pnl),
            summary="주식매도",
        )
        if loss:
            lines.append(loss)

    return lines


def trades_to_voucher_lines(
    trades: Iterable[Trade],
    sells: Iterable[SellResult],
    *,
    account_config: AccountConfig | None = None,
    partner_by_stock_id: Mapping[int, str] | None = None,
    remark_mode: str = "stock",
) -> list[VoucherLine]:
    """
    sells는 전체 이력 FIFO 결과여야 매도 원가가 정확하다.
    trades는 전표에 포함할 기간 필터된 거래.
    remark_mode: 'stock' | 'amount' — 전표 적요 형식.
    """
    mode = (remark_mode or "stock").strip().lower()
    if mode not in {"stock", "amount"}:
        mode = "stock"
    accounts = VoucherAccounts.from_config(account_config)
    sell_by_id = sells_by_trade_id(list(sells))
    ordered = sorted(trades, key=lambda t: (t.trade_date, t.id or 0))
    lines: list[VoucherLine] = []
    for trade in ordered:
        partner = _partner_for_trade(trade, partner_by_stock_id)
        if trade.side == "DIVIDEND":
            continue
        sell = None
        try:
            if trade.id is not None:
                sell = sell_by_id.get(int(trade.id))
        except (TypeError, ValueError):
            sell = None
        if trade.side == "BUY":
            lines.extend(
                _buy_lines(trade, accounts, partner, remark_mode=mode)
            )
        else:
            lines.extend(
                _sell_lines(trade, sell, accounts, partner, remark_mode=mode)
            )
    if mode == "amount":
        # '/' 기준 앞부분(통화·금액)만 최장 길이에 맞춰 공백 패딩
        _pad_summaries(lines)
    return lines


def count_voucher_lines(
    trades: Iterable[Trade],
    sells: Iterable[SellResult],
    *,
    account_config: AccountConfig | None = None,
    partner_by_stock_id: Mapping[int, str] | None = None,
    remark_mode: str = "stock",
) -> int:
    return len(
        trades_to_voucher_lines(
            trades,
            sells,
            account_config=account_config,
            partner_by_stock_id=partner_by_stock_id,
            remark_mode=remark_mode,
        )
    )


def _convert_xls_to_xlsx(xls_path: Path, xlsx_path: Path) -> None:
    """원본 .xls 템플릿을 openpyxl용 .xlsx로 1회 변환(값·병합 유지)."""
    import xlrd

    book = xlrd.open_workbook(str(xls_path), formatting_info=False)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet.name or "Sheet1"

    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            val = sheet.cell_value(r, c)
            if val == "":
                continue
            cell = ws.cell(row=r + 1, column=c + 1, value=val)
            if r + 1 == HEADER_ROW:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )

    for rlo, rhi, clo, chi in sheet.merged_cells:
        # xlrd: rhi/chi exclusive
        ws.merge_cells(
            start_row=rlo + 1,
            end_row=rhi,
            start_column=clo + 1,
            end_column=chi,
        )

    ws.row_dimensions[HEADER_ROW].height = 40
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _load_template_workbook() -> Workbook:
    """원본 양식(xlsx 캐시, 없으면 xls에서 변환)을 로드한다."""
    if not TEMPLATE_XLSX.exists():
        if not TEMPLATE_XLS.exists():
            raise FileNotFoundError(
                f"전표 양식 파일이 없습니다: {TEMPLATE_XLS}\n"
                "데스크톱의 '엑셀자료일반전표전송.xls'를 templates 폴더에 복사해 주세요."
            )
        _convert_xls_to_xlsx(TEMPLATE_XLS, TEMPLATE_XLSX)
    return load_workbook(TEMPLATE_XLSX)


def _clear_data_rows(ws) -> None:
    """템플릿 데이터 영역(11행~)을 비운다."""
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)


def _fill_company(ws, company_name: str = "", biz_reg_no: str = "") -> None:
    """원본 양식 위치: 회사명 L3(12), 사업자등록번호 O3(15)."""
    if company_name:
        ws.cell(row=3, column=12, value=company_name)
    if biz_reg_no:
        ws.cell(row=3, column=15, value=biz_reg_no)


def build_voucher_workbook(
    trades: list[Trade],
    sells: list[SellResult],
    *,
    company_name: str = "",
    biz_reg_no: str = "",
    account_config: AccountConfig | None = None,
    partner_by_stock_id: Mapping[int, str] | None = None,
    remark_mode: str = "stock",
) -> Workbook:
    wb = _load_template_workbook()
    ws = wb.active
    _clear_data_rows(ws)
    _fill_company(ws, company_name=company_name, biz_reg_no=biz_reg_no)

    lines = trades_to_voucher_lines(
        trades,
        sells,
        account_config=account_config,
        partner_by_stock_id=partner_by_stock_id,
        remark_mode=remark_mode,
    )
    for offset, line in enumerate(lines):
        row_idx = DATA_START_ROW + offset
        for col_idx, value in enumerate(line.as_row(), start=1):
            if value == "" or value is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 적요 열(H, index 8) 텍스트 서식 — 공백 패딩 유지
            if col_idx == 8 and isinstance(value, str):
                cell.number_format = "@"
    return wb


def export_voucher_excel_bytes(
    trades: list[Trade],
    sells: list[SellResult],
    *,
    company_name: str = "",
    biz_reg_no: str = "",
    account_config: AccountConfig | None = None,
    partner_by_stock_id: Mapping[int, str] | None = None,
    remark_mode: str = "stock",
) -> bytes:
    wb = build_voucher_workbook(
        trades,
        sells,
        company_name=company_name,
        biz_reg_no=biz_reg_no,
        account_config=account_config,
        partner_by_stock_id=partner_by_stock_id,
        remark_mode=remark_mode,
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
